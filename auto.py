from openpyxl import load_workbook
import random
import datetime

wb = load_workbook(filename = './Book1.xlsx')
sheet_names = wb.sheetnames
sheet = wb[sheet_names[0]]
copy_sheet = wb.copy_worksheet(sheet)
now = datetime.datetime.now()
copy_sheet.title = 'NewShift' + now.strftime(f'%Y%m%d%H%M%S')
original_lists = []
col_lists = copy_sheet.iter_cols(min_row=5, max_row=16, min_col=6, max_col=33)
for index, col in enumerate(col_lists):
  col_vals = []
  for i, cell in enumerate(col):
    col_vals.append(cell.value)
  original_lists.append(col_vals)
print(original_lists)

result_lists = []
for index, col in enumerate(original_lists):
  num = [1,2,3,4,5,6,7,8,9,10]
  random.shuffle(num)
  count = 0
  col_vals = []
  if index >= 1:
    while count < 12:
      for i, cell in enumerate(result_lists[index - 1]):
        if index == 1:
          if col[i] != None:
            count += 1
            col_vals.append(col[i])
          else:
            if cell == num[0]:
              num = [1,2,3,4,5,6,7,8,9,10]
              random.shuffle(num)
              count = 0
              col_vals = []
              break
            else:
              count += 1
              col_vals.append(num[0])
              num.pop(0)
        if index > 1:
          if col[i] != None:
            count += 1
            col_vals.append(col[i])
          else:
            if num[0] == cell or num[0] == result_lists[index - 2][i]:
              num = [1,2,3,4,5,6,7,8,9,10]
              random.shuffle(num)
              count = 0
              col_vals = []
              break
            else:
              count += 1
              col_vals.append(num[0])
              num.pop(0)
  else:
    for i, cell in enumerate(col):
      if len(num) > 0 and cell == None:
        col_vals.append(num[0])
        num.pop(0)
      else:
        col_vals.append(cell)
  result_lists.append(col_vals)

print('result_lists')
print(result_lists)

col_lists = copy_sheet.iter_cols(min_row=5, max_row=16, min_col=6, max_col=33)

for index, col in enumerate(col_lists):
  for i, cell in enumerate(col):
    cell.value = result_lists[index][i]

wb.save('Book1.xlsx')